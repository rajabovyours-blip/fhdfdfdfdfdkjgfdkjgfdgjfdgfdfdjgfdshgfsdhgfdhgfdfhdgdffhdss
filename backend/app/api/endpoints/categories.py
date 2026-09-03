from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List
from uuid import UUID

from app.db.session import get_db
from app.models.category import Category
from app.models.user import User
from app.schemas.product import CategoryModel, CategoryCreate, CategoryUpdate
from app.schemas.common import APIResponse
from app.api.dependencies import get_current_admin

router = APIRouter()

@router.get("", response_model=APIResponse[List[CategoryModel]])
async def get_categories(db: AsyncSession = Depends(get_db)):
    from sqlalchemy import select, func
    from app.models.product import Product

    # Subquery to count products per category
    count_stmt = (
        select(func.count(Product.id))
        .where(Product.category_id == Category.id)
        .scalar_subquery()
    )
    
    result = await db.execute(
        select(Category, count_stmt.label("product_count"))
    )
    
    response_data = []
    for category, p_count in result.all():
        category.product_count = p_count or 0
        response_data.append(CategoryModel.model_validate(category))
        
    return APIResponse(data=response_data)

@router.post("", response_model=APIResponse[CategoryModel])
async def create_category(category_in: CategoryCreate, db: AsyncSession = Depends(get_db), admin: User = Depends(get_current_admin)):
    from uuid import uuid4
    new_cat = Category(
        id=uuid4(),
        name=category_in.name,
        description=category_in.description or {"uz": "", "ru": "", "en": ""},
        icon_url=category_in.icon_url,
        image_url=category_in.image_url,
        parent_id=category_in.parent_id,
        is_featured=category_in.is_featured
    )
    db.add(new_cat)
    await db.commit()
    await db.refresh(new_cat)
    return APIResponse(data=CategoryModel.model_validate(new_cat))

@router.put("/{id}", response_model=APIResponse[CategoryModel])
async def update_category(id: UUID, category_in: CategoryUpdate, db: AsyncSession = Depends(get_db), admin: User = Depends(get_current_admin)):
    result = await db.execute(select(Category).where(Category.id == id))
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    update_data = category_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(category, field, value)
        
    await db.commit()
    await db.refresh(category)
    return APIResponse(data=CategoryModel.model_validate(category))

@router.delete("/{id}", response_model=APIResponse[dict])
async def delete_category(id: UUID, db: AsyncSession = Depends(get_db), admin: User = Depends(get_current_admin)):
    from sqlalchemy.orm import selectinload
    result = await db.execute(select(Category).where(Category.id == id).options(selectinload(Category.products)))
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
        
    if category.products:
        raise HTTPException(status_code=400, detail="Ushbu kategoriyada mahsulotlar mavjud. Avval mahsulotlarni o'chiring yoki boshqa kategoriyaga o'tkazing.")
    
    await db.delete(category)
    await db.commit()
    return APIResponse(data={"message": "Category deleted successfully"})

@router.get("/import/template")
async def download_categories_template(admin: User = Depends(get_current_admin)):
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kategoriyalar"
    
    headers = [
        "name_uz", "name_ru", "name_en",
        "description_uz", "description_ru", "description_en",
        "icon_url", "image_url", "is_featured (0 yoki 1)", "order_index"
    ]
    ws.append(headers)
    
    ws.append(["Elektronika", "Электроника", "Electronics", "Elektronika mahsulotlari", "Электронные товары", "Electronic goods", "", "", 1, 0])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=categories_template.xlsx"}
    )

@router.post("/import", response_model=APIResponse[dict])
async def import_categories(
    file: __import__('fastapi').UploadFile = __import__('fastapi').File(...), 
    db: AsyncSession = Depends(get_db), 
    admin: User = Depends(get_current_admin)
):
    import io
    import openpyxl
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Faqat .xlsx fayllar qabul qilinadi")
    
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Faylni o'qishda xatolik: {str(e)}")
        
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Fayl bo'sh yoki faqat sarlavha mavjud")
        
    headers = [str(h).strip() if h else "" for h in rows[0]]
    if "name_uz" not in headers:
         raise HTTPException(status_code=400, detail="Noto'g'ri shablon. 'name_uz' ustuni topilmadi.")
         
    from uuid import uuid4
    categories_to_add = []
    header_map = {col: i for i, col in enumerate(headers)}
    
    for i, row in enumerate(rows[1:], start=2):
        def get_val(key, default=""):
            if key not in header_map: return default
            idx = header_map[key]
            if idx >= len(row): return default
            val = row[idx]
            return str(val).strip() if val is not None else default
            
        name_uz = get_val("name_uz")
        if not name_uz:
            continue
            
        feat_val = get_val("is_featured (0 yoki 1)", "0")
        is_featured = feat_val == "1" or feat_val.lower() == "true"
        
        try:
            order_index = int(get_val("order_index", "0"))
        except:
            order_index = 0
            
        new_cat = Category(
            id=uuid4(),
            name={"uz": name_uz, "ru": get_val("name_ru"), "en": get_val("name_en")},
            description={"uz": get_val("description_uz"), "ru": get_val("description_ru"), "en": get_val("description_en")},
            icon_url=get_val("icon_url") or None,
            image_url=get_val("image_url") or None,
            is_featured=is_featured,
            order_index=order_index,
            parent_id=None
        )
        categories_to_add.append(new_cat)
        
    if not categories_to_add:
         raise HTTPException(status_code=400, detail="Yaroqli ma'lumotlar topilmadi.")
         
    db.add_all(categories_to_add)
    await db.commit()
    
    return APIResponse(data={"message": f"{len(categories_to_add)} ta kategoriya muvaffaqiyatli import qilindi."})

